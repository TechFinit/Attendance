from django.shortcuts import render

from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.contrib.auth import authenticate
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.timezone import now
from django.http import HttpResponse
import openpyxl

from openpyxl.styles import Font, Alignment
from .models import Attendance


# 🔹 LOGIN (OPTIMIZED + DUPLICATE PREVENTION + FAST RESPONSE)
@csrf_exempt
@api_view(['POST'])
def login_view(request):
    username = request.data.get('username')
    password = request.data.get('password')

    user = authenticate(username=username, password=password)

    if user:
        role = "tl" if user.is_staff else "employee"

        # ✅ EMPLOYEE FLOW
        if role == "employee":

            # 🔥 PREVENT DUPLICATE LOGIN
            existing = Attendance.objects.filter(
                user=user,
                logout_time__isnull=True
            ).last()

            if existing:
                return Response({
                    "error": "Already logged in! Please logout first."
                }, status=400)

            # ✅ CREATE ATTENDANCE
            attendance = Attendance.objects.create(user=user)

            # ✅ RETURN DATA DIRECTLY (NO NEED FOR /today/ CALL)
            return Response({
                "message": "Login successful",
                "role": role,
                "user": user.username,
                "login_time": attendance.login_time,
                "logout_time": None
            })

        # ✅ TL FLOW (NO ATTENDANCE ENTRY)
        return Response({
            "message": "Login successful",
            "role": role
        })

    return Response({"error": "Invalid credentials"}, status=400)


# 🔹 LOGOUT
@csrf_exempt
@api_view(['POST'])
def logout_view(request):
    username = request.data.get('username')

    attendance = Attendance.objects.filter(
        user__username=username,
        logout_time__isnull=True
    ).last()

    if attendance:
        attendance.logout_time = timezone.now()
        attendance.save()
        return Response({"message": "Logout successful"})
    
    # ✅ TL OR NO SESSION
    return Response({"message": "No active session or TL user"})


# 🔹 ATTENDANCE LIST (MONTH FILTER)
@api_view(['GET'])
def attendance_list(request):
    month = request.GET.get('month')

    records = Attendance.objects.all()

    if month:
        records = records.filter(login_time__month=int(month))

    data = []

    for record in records:
        if record.logout_time:
            duration = record.logout_time - record.login_time

            total_seconds = int(duration.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60

            total_hours = f"{hours}.{minutes:02d}"
        else:
            total_hours = "0.00"

        login_date = record.login_time.date()
        shift = "Night" if record.login_time.hour >= 18 else "Morning"

        data.append({
            "user": record.user.username,
            "date": login_date,
            "shift": shift,
            "login_time": record.login_time,
            "logout_time": record.logout_time,
            "total_hours": total_hours
        })

    return Response(data)


# 🔹 TODAY STATUS (STILL AVAILABLE IF NEEDED)
@api_view(['POST'])
def today_status(request):
    username = request.data.get('username')

    today = now().date()

    attendance = Attendance.objects.filter(
        user__username=username,
        login_time__date=today
    ).last()

    if attendance:
        return Response({
            "user": attendance.user.username,
            "login_time": attendance.login_time,
            "logout_time": attendance.logout_time
        })

    return Response({"message": "No record for today"})


# 🔥 DOWNLOAD EXCEL (WITH MONTH FILTER)
@api_view(['GET'])
def download_excel(request):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    month = request.GET.get('month')

    users = set(Attendance.objects.values_list('user__username', flat=True))

    for user in users:
        ws = wb.create_sheet(title=user)

        headers = ["Date", "Login", "Logout", "Total Hours"]
        ws.append(headers)

        # ✅ Header styling
        for col in ws[1]:
            col.font = Font(bold=True)
            col.alignment = Alignment(horizontal='center')

        records = Attendance.objects.filter(user__username=user)

        if month:
            records = records.filter(login_time__month=int(month))

        for record in records:
            if record.logout_time:
                duration = record.logout_time - record.login_time

                total_seconds = int(duration.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60

                total_hours = f"{hours}.{minutes:02d}"
            else:
                total_hours = "0.00"

            ws.append([
                str(record.login_time.date()),
                record.login_time.strftime("%H:%M"),
                record.logout_time.strftime("%H:%M") if record.logout_time else "-",
                total_hours
            ])

        # ✅ Auto width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=employee_attendance.xlsx'

    wb.save(response)
    return response